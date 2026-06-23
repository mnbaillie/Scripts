"""
SFR ZoneBudget v0.4

This script generates a ZoneBudget-style water budget of the surface water
network of a MODFLOW-OWHMv2 model. It takes as input the SFR input file (to
generate a network routing diagram), a user-supplied zone file, and the reach-
by-reach streamflow output file. It generates a spreadsheet with the surface
water budget for each zone each model timestep, accounting for the following
components:
    - FLOW_HEAD: Streamflow entering the zone at headwater segments
    - ADDITIONAL_INFLOW: Prescribed FLOW entering non-head segments (for example,
                         FLOW specified on a segment that also receives routed flow)
    - FLOW_SEEPAGE: Groundwater-surface water interaction (positive is loss to
                    groundwater to match groundwater ZoneBudget sign convention)
    - FLOW_OUTOFMODEL: Streamflow leaving the model domain from the zone
    - RUNOFF: Land surface runoff entering the stream network
    - FARM_DIVERSION_NET_QC: FMP diversions leaving the stream network
    - DIVERSION_INTERNAL_QC: SFR diversions that stay within the zone (not
                             included in the water budget calculation)
    - PRECIP: Direct precipitation into the stream network
    - STREAM_ET: Direct ET from the stream network
    - IN_FROM_ZONE_N: Inflow to this zone from zone N
    - OUT_TO_ZONE_N: Outflow from this zone to zone N
    - MASS_BALANCE_RESIDUAL: Difference between total inflows and outflows.
                             Because there is no storage in the stream network
                             in SFR, this should always be extremely small,
                             representing the effects of rounding of the
                             various water budget components. If the mass
                             balance residual is not insignificant, this
                             likely indicates an issue in the script.
The script generates a surface water budget for the entire stream network (tab
SFR_TOTAL), and automatically assigns any portion of the stream network not
given a zone number to zone 0.

Created by M. Baillie, West Yost, on 30 Dec 2025 with AI assistance. Please
contact mbaillie@westyost.com with questions, issues, and suggestions.

sfr_zonebudget_runfile.py

ZoneBudget-style surface-water balance for MODFLOW (MF-2005 lineage) SFR "DB" output.

Key features
------------
- Reads SFR input file (SFR2-style) to build a segment routing table (stress period 1),
  and writes a routing CSV for QC.
- Reads SFR output file:
    * ASCII table (whitespace- or comma-delimited), OR
    * Binary fixed-length record table like USGS "DB" binary (DATE_START + ints + doubles).
      (Zipped binary also supported.)
    * SFR ISTCB2 positive formatted reach-by-reach listing output.
    * Experimental full-table binary reach records used by the companion reader.
      SFR2 negative-ISTCB2 binary cell arrays are detected and rejected with a clear error because they are not full reach-by-reach budgets.
- Reads a zone configuration file with TWO options:
    * by-segment: Segment, Zone
    * by-reach: Segment, Reach, Zone
  Anything unspecified is assigned Zone = 0 (QC).
- Produces an Excel workbook:
    * README_METADATA tab
    * One tab per zone with a timeseries (one row per model timestep)
      and interzone exchange columns.
- Diversions:
    * Many SFR models report diversions as negative RUNOFF on the SOURCE reach.
    * RUNOFF at the source can be the sum of natural runoff (+) and diversion (-),
      so we compute a NET diversion indicator:
          DIVERSION_NET = max(0, -RUNOFF_reported)
    * We map diversion SOURCE->DESTINATION using SFR input (IUPSEG relationships),
      allocate DIVERSION_NET to destination segment(s), and treat cross-zone diversions
      as interzone transfers.
    * Internal (same-zone) net diversions are reported in DIVERSION_INTERNAL_QC.

IMPORTANT NOTES
---------------
- Diversion transfers are NET indicators (derived from negative RUNOFF), not guaranteed
  to equal "gross diversion". See README_METADATA in the output workbook.
- Interzone downstream transfers are computed at the SEGMENT level using segment outflow
  (FLOW_OUT at last reach). This is typically correct for routing boundaries; if you
  need reach-level boundary routing, we can extend using reach connectivity.

USER SETTINGS (EDIT THESE)
--------------------------
Set file paths below, then run this script (e.g., in Spyder) WITHOUT command-line args.
"""

from __future__ import annotations

import os
import re
import zipfile
import struct
from io import StringIO
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# =========================
# USER SETTINGS (EDIT ME)
# =========================
SFR_INPUT_PATH = r"Y:\mbaillie\SFRZB\Test Models\test2\test2.sfr"          # SFR input file (text)

# Optional LAK package input file (text). Leave blank if the model does not use LAK,
# or if you only want routing inferred from the SFR file. When supplied, the script
# parses NLAKES and connected sublake systems (NSLMS / IC / ISUB) and writes
# lake-aware routing QC outputs. This is important when one SFR segment flows into
# one sublake and another SFR segment receives outflow from a connected sublake.
LAK_INPUT_PATH = r"Y:\mbaillie\SFRZB\Test Models\test2\test2.lak"

# Optional LAK budget CSV produced by the companion listing-file scraper. Leave blank
# if the model does not use LAK, or if you only want SFR-only budgets. Expected
# columns include per, stp, lake, precip, evap, runoff, gw_inflow, gw_outflow,
# sw_inflow, sw_outflow, water_use, connected_lake_influx, and volume_change.
# Lake zones are assigned in the by-segment zone CSV using negative Segment values
# (for example, Segment=-3 assigns Lake 3 to a zone).
LAK_BUDGET_CSV_PATH = r"Y:\mbaillie\SFRZB\Test Models\test2\test2_lakbud.csv"

SFR_OUTPUT_PATH = r"Y:\mbaillie\SFRZB\Test Models\test2\test2_rbr.csv"  # ASCII or binary DB output; can be .zip
ZONE_CONFIG_PATH = r"Y:\mbaillie\SFRZB\Test Models\test2\test2_zone_all1zone.csv"       # by-segment or by-reach
OUT_EXCEL_PATH = r"Y:\mbaillie\SFRZB\Test Models\test2\test2_SFRZB_all1zone.xlsx"

# Optional: choose where to write routing CSV (QC). If blank, writes next to OUT_EXCEL_PATH.
OUT_ROUTING_CSV_PATH = r"Y:\mbaillie\SFRZB\Test Models\test2\test2_all1zone_RoutingQC.csv"  # e.g. r"Y:\path\to\routing.csv"

# Optional: choose where to write a lake-aware routing edge list (QC). If blank, writes
# next to OUT_ROUTING_CSV_PATH using the suffix _edges.csv. This file includes
# stream-to-stream, stream-to-lake, lake-to-stream, and LAK sublake-system edges.
OUT_ROUTING_EDGES_CSV_PATH = r""

# If your binary output is zipped and contains multiple files, set this to the member name.
# If blank, the script will use the first member in the zip.
ZIP_MEMBER_NAME = r""

# Advanced SFR output reader options. These are used only for binary reach-by-reach
# records that are neither DB binary nor SFR2 negative-ISTCB2 cell-array output.
BINARY_RBR_NFLOAT = 11        # 11 if no streambed elevation field, 12 if included
BINARY_RBR_PRECISION = "auto" # "auto", "double", or "single"
BINARY_RBR_ENDIAN = "<"       # "<" little-endian, ">" big-endian
BINARY_RBR_FORTRAN = False    # True if each binary reach row has Fortran record markers

# Diagnostic option only. Keep False for ZoneBudget work. When True, SFR2
# negative-ISTCB2 cell-array files are returned as gridded cell values instead
# of raising an error. They are not usable for the reach-by-reach budget.
ALLOW_SFR2_NEGATIVE_ISTCB2_CELL_ARRAY = False

# Residual-flow tolerance for separating routed inflow from additional prescribed
# inflow at non-head segments. Values at or below this threshold are treated as
# numerical noise/roundoff.
ADDITIONAL_INFLOW_TOLERANCE = 1.0e-9

# Unit conversion and output basis
# MODFLOW SFR output fluxes are in MODEL_VOLUME / MODEL_TIME.
# This script can report either:
#   - integrated volume per timestep ("PER_STRESS_PERIOD"), or
#   - average rate as volume-per-day ("PER_DAY").
#
# Volume units
# ------------
# Choose a MODEL_LENGTH_UNIT (what the model uses internally) and an OUTPUT_VOLUME_UNIT.
# Common OUTPUT_VOLUME_UNIT options include "ft3", "m3", and "acft" (acre-feet).
# If you prefer, set OUTPUT_VOLUME_UNIT="custom" and provide VOLUME_CONV_FACTOR.
#
# Notes:
# - MODFLOW outputs are always in cubic length units (e.g., ft^3 or m^3), not mixed units like acre-feet.
# - Conversion to acre-feet is therefore one-way (cubic length -> acre-feet).

MODEL_LENGTH_UNIT = "ft"          # "ft", "m", or "custom"
OUTPUT_VOLUME_UNIT = "ft3"        # "ft3", "m3", "acft", or "custom"

# If MODEL_LENGTH_UNIT is "custom", provide the number of feet per model length unit.
# Example: inches -> FT_PER_MODEL_LEN = 1/12
FT_PER_MODEL_LEN = 1.0

# If OUTPUT_VOLUME_UNIT is "custom", provide conversion from model volume units to desired output volume units:
#   output_volume = model_volume * VOLUME_CONV_FACTOR
VOLUME_CONV_FACTOR = 1.0

# Time units
# ----------
# MODEL_TIME_UNIT_IN_DAYS: length of one model time unit in days.
#   e.g., model time unit is days  -> 1.0
#         model time unit is hours -> 1.0 / 24.0
#         model time unit is years -> 365.25
MODEL_TIME_UNIT_IN_DAYS = 1.0

# OUTPUT_BASIS controls how fluxes are reported in Excel:
#   "PER_STRESS_PERIOD" -> integrated volume over each timestep (volume per stress period)
#   "PER_DAY"           -> average rate over the timestep (volume per day)
OUTPUT_BASIS = "PER_DAY"  # or "PER_STRESS_PERIOD"

# Optional label for the output workbook metadata.
VOLUME_UNIT_LABEL = "ft^3"  # e.g., "ac-ft", "m^3"


# =========================
# INTERNALS (no edits needed)
# =========================
INT_PAT = re.compile(r'^[+-]?\d+$')
FLOAT_PAT = re.compile(r'^[+-]?(\d+(\.\d*)?|\.\d+)([Ee][+-]?\d+)?$')


def _strip_inline_comments(line: str) -> str:
    for sep in ("#", ";", "!"):
        if sep in line:
            line = line.split(sep, 1)[0]
    # allow comma-delimited numeric fields in SFR input
    line = line.replace(",", " ")
    return line.rstrip("\n")


def _is_comment_line(line: str) -> bool:
    s = line.strip()
    if not s:
        return True
    if s.startswith(("#", ";", "!")):
        return True
    if line and line[0] in ("C", "c", "*"):
        return True
    return False


def _tokens(line: str) -> List[str]:
    return _strip_inline_comments(line).strip().split()


def infer_counts_sfr(lines: List[str], max_lines: int = 50000) -> Tuple[int, int, int, str]:
    # 1) annotated line containing NSTRM
    for i, line in enumerate(lines[:max_lines]):
        if "NSTRM" in line.upper() and not line.strip().startswith(("#", ";", "!", "C", "c", "*")):
            tk = _tokens(line)
            if len(tk) >= 2 and INT_PAT.match(tk[0]) and INT_PAT.match(tk[1]):
                return i, int(tk[0]), int(tk[1]), line.strip()

    # 2) fallback: plausible dataset 1c
    best = None
    for i, line in enumerate(lines[:max_lines]):
        if _is_comment_line(line):
            continue
        tk = _tokens(line)
        if len(tk) < 5:
            continue
        if not (INT_PAT.match(tk[0]) and INT_PAT.match(tk[1])):
            continue
        nstrm = int(tk[0])
        nss = int(tk[1])
        if nss <= 0 or abs(nstrm) < nss or nss > 500000 or abs(nstrm) > 50000000:
            continue
        if not all(FLOAT_PAT.match(t) or INT_PAT.match(t) for t in tk[: min(len(tk), 12)]):
            continue
        score = len(tk) + nss / 1000.0
        if best is None or score > best[0]:
            best = (score, i, nstrm, nss, line.strip())

    if best is None:
        raise ValueError("Could not infer NSTRM/NSS from the SFR input.")
    _, i, nstrm, nss, raw = best
    return i, nstrm, nss, raw


def _looks_like_segment_header(tk: List[str], expected_seg: int) -> bool:
    if len(tk) < 4:
        return False
    if not all(INT_PAT.match(x) for x in tk[:4]):
        return False
    seg = int(tk[0])
    icalc = int(tk[1])
    if seg != expected_seg:
        return False
    if icalc not in (0, 1, 2, 3, 4):
        return False
    return True


def _detect_reachinput(lines: List[str], search_lines: int = 100) -> bool:
    return any("REACHINPUT" in l.upper() for l in lines[:search_lines])


def _find_first_reach_record_line(lines: List[str], start_idx: int, lookahead: int = 500) -> Optional[int]:
    """
    Reach records (dataset 2) typically begin with at least 5 integers: K I J SEG RCH ...
    Search forward for the first non-comment line matching that shape.
    """
    for j in range(start_idx, min(len(lines), start_idx + lookahead)):
        if _is_comment_line(lines[j]):
            continue
        tk = _tokens(lines[j])
        if len(tk) >= 5 and all(INT_PAT.match(x) for x in tk[:5]):
            return j
    return None


def find_segment_block_start(lines: List[str], nstrm: int, counts_idx: int) -> int:
    """
    Locate the first stress-period segment block start.

    QC fix:
      - Many SFR files have the reach list immediately after the NSTRM/NSS line even without the
        REACHINPUT keyword. To avoid false positives, ALWAYS skip abs(NSTRM) reach records
        after the counts line (starting at the first reach-like record) before searching for segment headers.
    """
    scan_start = 0

    reach_start = _find_first_reach_record_line(lines, counts_idx + 1)
    if reach_start is not None:
        reach_count = 0
        j = reach_start
        while j < len(lines) and reach_count < abs(int(nstrm)):
            if _is_comment_line(lines[j]):
                j += 1
                continue
            tk = _tokens(lines[j])
            if len(tk) >= 5 and all(INT_PAT.match(x) for x in tk[:5]):
                reach_count += 1
            j += 1
        scan_start = max(scan_start, j)

    for i in range(scan_start, len(lines)):
        if _is_comment_line(lines[i]):
            continue
        tk = _tokens(lines[i])
        if _looks_like_segment_header(tk, 1):
            nonc = 0
            for j in range(i + 1, min(len(lines), i + 12000)):
                if _is_comment_line(lines[j]):
                    continue
                nonc += 1
                tk2 = _tokens(lines[j])
                if _looks_like_segment_header(tk2, 2):
                    return i
                if nonc > 2000:
                    break
    raise ValueError("Could not locate the start of the segment data block (stress period 1).")


def parse_sfr_routing_table(sfr_input_path: str) -> Dict:
    with open(sfr_input_path, "r", encoding="utf-8", errors="ignore") as f:
        lines = f.readlines()

    idx_counts, nstrm, nss, counts_raw = infer_counts_sfr(lines)
    start = find_segment_block_start(lines, nstrm=nstrm, counts_idx=idx_counts)

    rows = []
    expected = 1
    i = start
    while expected <= nss and i < len(lines):
        if _is_comment_line(lines[i]):
            i += 1
            continue
        tk = _tokens(lines[i])
        if _looks_like_segment_header(tk, expected):
            seg = int(tk[0])
            icalc = int(tk[1])
            outseg = int(tk[2])
            iupseg = int(tk[3])
            rows.append(
                dict(
                    segment=seg,
                    icalc=icalc,
                    outseg=outseg,
                    iupseg=iupseg,
                    header_line_no=i + 1,
                    header_raw=_strip_inline_comments(lines[i]).strip(),
                )
            )
            expected += 1
        i += 1

    if expected <= nss:
        raise ValueError(f"Parsed only segments 1..{expected-1} of NSS={nss}.")

    rt = pd.DataFrame(rows)

    rt["outseg_norm"] = rt["outseg"].where(rt["outseg"].between(1, nss), 0).astype(int)
    rt["iupseg_norm"] = rt["iupseg"].where(rt["iupseg"].between(1, nss), 0).astype(int)
    rt["is_out_of_model"] = ~rt["outseg"].between(1, nss)
    rt["is_diversion_segment"] = rt["iupseg"].between(1, nss)

    inflow_targets = set(rt.loc[rt["outseg_norm"] > 0, "outseg_norm"].astype(int)).union(
        set(rt.loc[rt["iupseg_norm"] > 0, "segment"].astype(int))
    )
    rt["is_head_segment"] = ~rt["segment"].astype(int).isin(inflow_targets)

    return dict(
        nstrm=nstrm,
        nss=nss,
        counts_line_no=idx_counts + 1,
        counts_line=counts_raw,
        segblock_start_line_no=start + 1,
        routing_table=rt,
    )



def parse_lak_sublake_systems(lak_input_path: str) -> Dict[str, Any]:
    """Parse basic LAK package information needed for routing QC.

    This intentionally reads only the parts needed for topology:
      - NLAKES from Data Set 1
      - NSLMS from Data Set 7
      - IC and ISUB(1)..ISUB(IC) from Data Set 8a, repeated NSLMS times

    LAK files vary in how much annotation they include. The parser first looks
    for lines annotated with NLAKES and NSLMS. If NSLMS is not annotated, the
    parser returns NLAKES and a warning rather than guessing aggressively.
    """
    if not lak_input_path or not str(lak_input_path).strip():
        return {
            "lak_input": "",
            "nlakes": 0,
            "nslms": 0,
            "sublake_systems": [],
            "warnings": ["No LAK input file supplied; connected sublake systems cannot be represented."],
        }

    with open(lak_input_path, "r", encoding="utf-8", errors="ignore") as f:
        raw_lines = f.readlines()

    clean = []
    for idx, line in enumerate(raw_lines):
        if _is_comment_line(line):
            continue
        tk = _tokens(line)
        if tk:
            clean.append((idx + 1, line.rstrip("\n"), tk))

    warnings: List[str] = []
    nlakes = 0
    nslms = 0
    systems: List[Dict[str, Any]] = []

    # NLAKES is the first integer on Data Set 1. Prefer annotated line, then
    # first non-comment integer line as a fallback.
    nlakes_idx = None
    for i, (lineno, raw, tk) in enumerate(clean):
        if "NLAKES" in raw.upper() and INT_PAT.match(tk[0]):
            nlakes = int(tk[0])
            nlakes_idx = i
            break
    if nlakes_idx is None and clean and INT_PAT.match(clean[0][2][0]):
        nlakes = int(clean[0][2][0])
        nlakes_idx = 0
    if nlakes <= 0:
        warnings.append("Could not parse a positive NLAKES value from the LAK input file.")

    # NSLMS is Data Set 7. Prefer annotation to avoid confusing it with other
    # LAK data sets. If absent, do not guess because false positives would be
    # worse than omitting lake-to-lake QC edges.
    nslms_idx = None
    for i, (lineno, raw, tk) in enumerate(clean):
        if "NSLMS" in raw.upper() and INT_PAT.match(tk[0]):
            nslms = int(tk[0])
            nslms_idx = i
            break
    if nslms_idx is None:
        warnings.append("Could not locate an annotated NSLMS line in the LAK file; sublake systems were not parsed.")
        return {
            "lak_input": lak_input_path,
            "nlakes": nlakes,
            "nslms": 0,
            "sublake_systems": [],
            "warnings": warnings,
        }

    # Data Set 8a: one line/list per sublake system. IC can be followed by the
    # complete ISUB list on the same line or continued onto following lines.
    cursor = nslms_idx + 1
    for sys_no in range(1, nslms + 1):
        vals: List[int] = []
        start_line = None
        while cursor < len(clean):
            lineno, raw, tk = clean[cursor]
            cursor += 1
            ints = []
            for t in tk:
                if INT_PAT.match(t):
                    ints.append(int(t))
                else:
                    break
            if not ints:
                continue
            if start_line is None:
                start_line = lineno
            vals.extend(ints)
            if vals:
                ic = vals[0]
                if ic <= 0:
                    warnings.append(f"Sublake system {sys_no} has non-positive IC on line {start_line}; skipped.")
                    vals = []
                    break
                if len(vals) >= ic + 1:
                    lakes = vals[1:ic + 1]
                    systems.append({
                        "system_id": sys_no,
                        "ic": ic,
                        "sublakes": lakes,
                        "line_no": start_line,
                    })
                    break
        if len(systems) < sys_no:
            warnings.append(f"Could not parse complete IC/ISUB list for sublake system {sys_no}.")
            break

    return {
        "lak_input": lak_input_path,
        "nlakes": nlakes,
        "nslms": nslms,
        "sublake_systems": systems,
        "warnings": warnings,
    }


def build_routing_edges_table(routing: pd.DataFrame, lak_info: Optional[Dict[str, Any]] = None) -> pd.DataFrame:
    """Build a node-edge routing table for QC/visualization support.

    Node IDs are strings so lakes and stream segments remain explicit:
      - SFR segment 12 -> "SFR_12"
      - Lake 3        -> "LAKE_3"
    """
    rows: List[Dict[str, Any]] = []
    rt = routing.copy()
    for r in rt.itertuples(index=False):
        seg = int(r.segment)
        outseg = int(r.outseg)
        iupseg = int(r.iupseg)
        if outseg > 0:
            rows.append(dict(from_node=f"SFR_{seg}", to_node=f"SFR_{outseg}", from_type="SFR", to_type="SFR", edge_type="SFR_OUTSEG", from_id=seg, to_id=outseg))
        elif outseg < 0:
            lake = abs(outseg)
            rows.append(dict(from_node=f"SFR_{seg}", to_node=f"LAKE_{lake}", from_type="SFR", to_type="LAKE", edge_type="SFR_TO_LAKE_OUTSEG", from_id=seg, to_id=lake))
        if iupseg < 0:
            lake = abs(iupseg)
            rows.append(dict(from_node=f"LAKE_{lake}", to_node=f"SFR_{seg}", from_type="LAKE", to_type="SFR", edge_type="LAKE_TO_SFR_IUPSEG", from_id=lake, to_id=seg))

    if lak_info:
        for sys in lak_info.get("sublake_systems", []):
            lakes = [int(x) for x in sys.get("sublakes", [])]
            if len(lakes) < 2:
                continue
            anchor = lakes[0]
            for lake in lakes[1:]:
                rows.append(dict(
                    from_node=f"LAKE_{anchor}",
                    to_node=f"LAKE_{lake}",
                    from_type="LAKE",
                    to_type="LAKE",
                    edge_type="LAK_SUBLAKE_SYSTEM",
                    from_id=anchor,
                    to_id=lake,
                    sublake_system_id=int(sys.get("system_id", 0)),
                ))

    if not rows:
        return pd.DataFrame(columns=["from_node","to_node","from_type","to_type","edge_type","from_id","to_id","sublake_system_id"])
    edges = pd.DataFrame(rows)
    if "sublake_system_id" not in edges.columns:
        edges["sublake_system_id"] = np.nan
    return edges[["from_node","to_node","from_type","to_type","edge_type","from_id","to_id","sublake_system_id"]]


def build_lake_routing_qc(routing: pd.DataFrame, lak_info: Optional[Dict[str, Any]] = None) -> pd.DataFrame:
    """Flag potentially incomplete lake routing if LAK sublake systems are absent."""
    rows: List[Dict[str, Any]] = []
    rt = routing.copy()
    stream_to_lake = sorted(set(abs(int(x)) for x in rt.loc[rt["outseg"] < 0, "outseg"].tolist()))
    lake_to_stream = sorted(set(abs(int(x)) for x in rt.loc[rt["iupseg"] < 0, "iupseg"].tolist()))
    all_sfr_lakes = sorted(set(stream_to_lake) | set(lake_to_stream))

    systems = (lak_info or {}).get("sublake_systems", []) if lak_info else []
    lake_to_systems: Dict[int, List[int]] = {}
    for sys in systems:
        sid = int(sys.get("system_id", 0))
        for lake in sys.get("sublakes", []):
            lake_to_systems.setdefault(int(lake), []).append(sid)

    for lake in all_sfr_lakes:
        rows.append(dict(
            qc_type="SFR_LAKE_REFERENCE",
            lake=lake,
            related_sublake_systems=",".join(str(x) for x in lake_to_systems.get(lake, [])),
            message="Lake referenced by SFR OUTSEG/IUPSEG. Connected sublake context is available." if lake in lake_to_systems else "Lake referenced by SFR OUTSEG/IUPSEG. No connected sublake-system context was parsed for this lake.",
        ))

    if not systems and all_sfr_lakes:
        rows.append(dict(
            qc_type="LAK_FILE_RECOMMENDED",
            lake="",
            related_sublake_systems="",
            message="SFR references one or more lakes, but no connected sublake systems were parsed. Supply LAK_INPUT_PATH for models with connected sublakes.",
        ))

    for sys in systems:
        lakes = set(int(x) for x in sys.get("sublakes", []))
        inflow_lakes = sorted(lakes & set(stream_to_lake))
        outflow_lakes = sorted(lakes & set(lake_to_stream))
        if inflow_lakes and outflow_lakes and set(inflow_lakes) != set(outflow_lakes):
            rows.append(dict(
                qc_type="SUBLAKE_SYSTEM_CROSSES_SFR_CONNECTIONS",
                lake=";".join(str(x) for x in sorted(lakes)),
                related_sublake_systems=str(sys.get("system_id", "")),
                message=f"SFR inflow lake(s) {inflow_lakes} and SFR outflow lake(s) {outflow_lakes} occur within the same connected sublake system; SFR-only routing would not show the full connection.",
            ))

    return pd.DataFrame(rows, columns=["qc_type","lake","related_sublake_systems","message"])

def read_zone_config(zone_path: str) -> Tuple[str, pd.DataFrame]:
    """
    Read zone configuration.

    Supported:
      - by-segment: Segment, Zone
      - by-reach:   Segment, Reach, Zone

    Notes:
      - Column names are case-insensitive and whitespace-insensitive.
      - Blank lines and rows with missing required values are dropped.
      - Anything not specified later receives Zone=0 (QC).
    """
    # CSV first, else whitespace-delimited
    try:
        z = pd.read_csv(zone_path, comment="#")
    except Exception:
        z = pd.read_table(zone_path, sep=r"\s+", engine="python", comment="#")

    # normalize column names
    z.columns = [str(c).strip().lower() for c in z.columns]

    # drop fully empty rows
    z = z.dropna(how="all").copy()

    def _coerce_int(col: str) -> None:
        z[col] = pd.to_numeric(z[col], errors="coerce")

    if "reach" in z.columns:
        required = ["segment", "reach", "zone"]
        missing = [c for c in required if c not in z.columns]
        if missing:
            raise ValueError(f"By-reach zone file must have columns: Segment, Reach, Zone (missing: {missing})")

        for c in required:
            _coerce_int(c)

        z = z.dropna(subset=required).copy()
        if len(z) == 0:
            raise ValueError("Zone config has no valid rows after parsing (check headers/blank lines).")

        z["segment"] = z["segment"].astype(int)
        z["reach"] = z["reach"].astype(int)
        z["zone"] = z["zone"].astype(int)

        if (z["segment"] == 0).any() or (z["reach"] <= 0).any():
            raise ValueError("Zone config contains Segment=0 or non-positive Reach values. Use negative Segment values only for lake-zone rows in by-segment files.")
        if (z["segment"] < 0).any():
            raise ValueError("Lake-zone rows with negative Segment values are supported only for by-segment zone files, not by-reach files.")
        return "reach", z[required].copy()

    else:
        required = ["segment", "zone"]
        missing = [c for c in required if c not in z.columns]
        if missing:
            raise ValueError(f"By-segment zone file must have columns: Segment, Zone (missing: {missing})")

        for c in required:
            _coerce_int(c)

        z = z.dropna(subset=required).copy()
        if len(z) == 0:
            raise ValueError("Zone config has no valid rows after parsing (check headers/blank lines).")

        z["segment"] = z["segment"].astype(int)
        z["zone"] = z["zone"].astype(int)


        if (z["segment"] == 0).any():
            raise ValueError("Zone config contains Segment=0. Use positive Segment values for SFR segments and negative Segment values for lakes.")
        return "segment", z[required].copy()


def read_lake_budget_csv(path: str) -> Optional[pd.DataFrame]:
    """Read normalized LAK-budget CSV from the companion listing-file scraper.

    The LAK scraper is intentionally kept separate from this ZoneBudget script.
    This reader only consumes its CSV output and normalizes column names/types so
    lake budget terms can be included by zone. Leave path blank to disable lake
    budget handling.
    """
    if not path or not str(path).strip():
        return None
    if not os.path.exists(path):
        raise FileNotFoundError(f"LAK budget CSV was not found: {path}")

    lak = pd.read_csv(path, comment="#")
    lak.columns = [str(c).strip().lower() for c in lak.columns]

    required = ["per", "stp", "lake"]
    missing = [c for c in required if c not in lak.columns]
    if missing:
        raise ValueError(f"LAK budget CSV is missing required columns: {missing}")

    optional_numeric = [
        "delt", "pertim", "totim", "stage", "volume", "volume_change", "updated_volume",
        "precip", "evap", "runoff", "gw_inflow", "gw_outflow", "sw_inflow", "sw_outflow",
        "water_use", "connected_lake_influx", "surface_area", "percent_discrepancy",
    ]
    for c in required + optional_numeric:
        if c in lak.columns:
            lak[c] = pd.to_numeric(lak[c], errors="coerce")

    lak = lak.dropna(subset=["per", "stp", "lake"]).copy()
    if lak.empty:
        raise ValueError("LAK budget CSV has no valid rows after parsing PER/STP/LAKE.")

    lak["PER"] = lak["per"].astype(int)
    lak["STP"] = lak["stp"].astype(int)
    lak["LAKE"] = lak["lake"].astype(int)

    for src, dst in [
        ("precip", "LAK_PRECIP"),
        ("evap", "LAK_EVAP"),
        ("runoff", "LAK_RUNOFF"),
        ("gw_inflow", "LAK_GW_INFLOW"),
        ("gw_outflow", "LAK_GW_OUTFLOW"),
        ("sw_inflow", "LAK_SW_INFLOW_QC"),
        ("sw_outflow", "LAK_SW_OUTFLOW_QC"),
        ("water_use", "LAK_WATER_USE"),
        ("connected_lake_influx", "LAK_CONNECTED_LAKE_INFLUX_QC"),
        ("volume_change", "LAK_STORAGE_CHANGE"),
    ]:
        lak[dst] = lak[src].fillna(0.0).astype(float) if src in lak.columns else 0.0

    keep = [
        "PER", "STP", "LAKE", "LAK_PRECIP", "LAK_EVAP", "LAK_RUNOFF", "LAK_GW_INFLOW",
        "LAK_GW_OUTFLOW", "LAK_SW_INFLOW_QC", "LAK_SW_OUTFLOW_QC", "LAK_WATER_USE",
        "LAK_CONNECTED_LAKE_INFLUX_QC", "LAK_STORAGE_CHANGE",
    ]
    return lak[keep]


# =========================
# SFR OUTPUT READER SUPPORT
# =========================
LISTING_COLS_11 = [
    'LAYER','ROW','COL','SEG','RCH',
    'FLOW_IN','FLOW_SEEPAGE','FLOW_OUT','RUNOFF','PRECIP','STREAM_ET',
    'HEAD_STREAM','DEPTH_STREAM','WIDTH_STREAM','COND_STREAM','HEAD_GRADIENT'
]
LISTING_COLS_12 = LISTING_COLS_11 + ['ELEV_UP_STREAM']
DB_TAIL_COLS = [
    'FLOW_IN','FLOW_SEEPAGE','FLOW_OUT','RUNOFF','PRECIP','STREAM_ET',
    'HEAD_STREAM','HEAD_AQUIFER','DEPTH_STREAM','WIDTH_STREAM','LENGTH_STREAM',
    'HEAD_GRADIENT','COND_STREAM','ELEV_UP_STREAM'
]
REQ = ['PER','STP','SEG','RCH','FLOW_IN','FLOW_SEEPAGE','FLOW_OUT','RUNOFF','PRECIP','STREAM_ET']

_HEADER_RE = re.compile(r"STREAM\s+LISTING\s+PERIOD\s+(\d+)\s+STEP\s+(\d+)", re.I)
_NUM_RE = re.compile(r"^[+-]?(?:\d+(?:\.\d*)?|\.\d+)(?:[EeDd][+-]?\d+)?$")


class SfrOutputNotZoneBudgetReady(ValueError):
    """Raised when a recognized SFR output lacks the full reach-budget fields."""



def _looks_text(raw: bytes) -> bool:
    sample = raw[:4096]
    if not sample:
        return True
    nul = sample.count(b'\x00') / max(1, len(sample))
    if nul > 0.01:
        return False
    try:
        sample.decode('utf-8')
        return True
    except UnicodeDecodeError:
        try:
            sample.decode('cp1252')
            return True
        except Exception:
            return False


def _decode(raw: bytes) -> str:
    try:
        return raw.decode('utf-8')
    except UnicodeDecodeError:
        return raw.decode('cp1252', errors='ignore')


def read_sfr_reachbyreach_ascii_listing(text: str) -> pd.DataFrame:
    """Parse formatted SFR ISTCB2 reach-by-reach listing output.

    Handles the 11-float variant (no streambed elevation) and 12-float variant
    (with ELEVATION). Adds PER/STP from each block header. DELT and SIMTIME are
    not present in this listing format, so they are left blank/NaN rather than
    filled with synthetic placeholder values.
    """
    rows = []
    per: Optional[int] = None
    stp: Optional[int] = None
    tindex = 0
    last_key = None

    for line in text.splitlines():
        m = _HEADER_RE.search(line)
        if m:
            per, stp = int(m.group(1)), int(m.group(2))
            key = (per, stp)
            if key != last_key:
                tindex += 1
                last_key = key
            continue
        if per is None:
            continue
        toks = line.strip().replace('D','E').replace('d','E').split()
        if len(toks) not in (16, 17):
            continue
        if not all(_NUM_RE.match(t) for t in toks):
            continue
        try:
            ints = [int(float(toks[i])) for i in range(5)]
        except Exception:
            continue
        vals = [float(x) for x in toks[5:]]
        cols = LISTING_COLS_11 if len(toks) == 16 else LISTING_COLS_12
        rec = dict(zip(cols[:5], ints))
        rec.update(dict(zip(cols[5:], vals)))
        rec.update(PER=per, STP=stp, DELT=np.nan, SIMTIME=np.nan, DATE_START='', DATE_TIME=pd.NaT)
        rows.append(rec)

    if not rows:
        raise ValueError('No SFR reach-by-reach listing rows were found.')
    df = pd.DataFrame(rows)
    if 'ELEV_UP_STREAM' not in df.columns:
        df['ELEV_UP_STREAM'] = np.nan
    if 'HEAD_AQUIFER' not in df.columns:
        df['HEAD_AQUIFER'] = np.nan
    if 'LENGTH_STREAM' not in df.columns:
        df['LENGTH_STREAM'] = np.nan
    return _normalize(df, source_format='ascii_rbr_listing')


def read_db_ascii(text: str) -> pd.DataFrame:
    df = pd.read_csv(StringIO(text), sep=r'\s+|,', engine='python')
    if not set(['DATE_START','PER','STP','SEG','RCH']).issubset(df.columns):
        raise ValueError('Not a DB-style ASCII table')
    return _normalize(df, source_format='ascii_db')


def read_binary_db(raw: bytes) -> pd.DataFrame:
    """Read OWHM DB binary records: DATE_START, PER/STP/DELT/SIMTIME/SEG/RCH, doubles."""
    try:
        s19 = raw[:19].decode('ascii')
    except Exception:
        s19 = ''
    if not re.match(r"^\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}$", s19):
        raise ValueError('Not a DATE_START DB binary file')
    ts0 = raw[:19]
    rec_size = raw.find(ts0, 19)
    if rec_size <= 0:
        raise ValueError('Could not infer DB binary record size')
    hdr = struct.Struct('<19s i i d d i i')
    tail_bytes = rec_size - hdr.size
    if tail_bytes < 0 or tail_bytes % 8:
        raise ValueError('Invalid DB binary record size')
    n_tail = tail_bytes // 8
    tail_struct = struct.Struct('<' + 'd'*n_tail)
    nrec = len(raw) // rec_size
    recs = []
    for i in range(nrec):
        chunk = raw[i*rec_size:(i+1)*rec_size]
        ds, per, stp, delt, simtime, seg, rch = hdr.unpack(chunk[:hdr.size])
        tail = tail_struct.unpack(chunk[hdr.size:])
        rec = {'DATE_START': ds.decode('ascii','ignore'), 'PER':per, 'STP':stp, 'DELT':delt, 'SIMTIME':simtime, 'SEG':seg, 'RCH':rch}
        for c, v in zip(DB_TAIL_COLS, tail):
            rec[c] = v
        recs.append(rec)
    return _normalize(pd.DataFrame(recs), source_format='binary_db')


def read_sfr2_istcb2_negative_binary_array(raw: bytes) -> Tuple[Dict[str, Any], pd.DataFrame]:
    """Parse SFR2 ISTCB2<0 MODFLOW-style binary cell array records.

    Observed layout, per record, is the standard single-precision MODFLOW
    cell-by-cell array header:
        KSTP, KPER, TEXT(16), NCOL, NROW, NLAY, array(NCOL*NROW*abs(NLAY))

    Returns (metadata, dataframe). The dataframe is cell-based, with LAYER, ROW,
    COL, and FLOW_OUT for records labelled STREAMFLOW OUT. It does not contain
    SEG/RCH or the other reach-budget terms needed for ZoneBudget.
    """
    off = 0
    records = []
    headers = []
    hdr = struct.Struct('<ii16siii')
    while off < len(raw):
        if off + hdr.size > len(raw):
            raise ValueError('Trailing bytes before a complete MODFLOW array header')
        kstp, kper, text_b, ncol, nrow, nlay = hdr.unpack(raw[off:off+hdr.size])
        off += hdr.size
        text = text_b.decode('ascii', errors='ignore').strip()
        if ncol <= 0 or nrow <= 0 or nlay == 0 or abs(nlay) > 100000 or ncol > 100000 or nrow > 100000:
            raise ValueError('Not a plausible MODFLOW binary array header')
        nvals = ncol * nrow * abs(nlay)
        nbytes = nvals * 4
        if off + nbytes > len(raw):
            raise ValueError('MODFLOW binary array record is incomplete')
        vals = np.frombuffer(raw, dtype='<f4', count=nvals, offset=off).astype(float)
        off += nbytes
        headers.append(dict(KSTP=kstp, KPER=kper, TEXT=text, NCOL=ncol, NROW=nrow, NLAY=nlay, NVALS=nvals))
        arr = vals.reshape((abs(nlay), nrow, ncol))
        for ilay in range(abs(nlay)):
            layer_arr = arr[ilay]
            rr, cc = np.indices((nrow, ncol))
            d = pd.DataFrame({
                'PER': kper,
                'STP': kstp,
                'LAYER': ilay + 1,
                'ROW': rr.ravel() + 1,
                'COL': cc.ravel() + 1,
                'BINARY_TEXT': text,
                'FLOW_OUT': layer_arr.ravel(),
            })
            records.append(d)
    if not records:
        raise ValueError('No MODFLOW binary array records found')
    df = pd.concat(records, ignore_index=True)
    df['DELT'] = 1.0
    df['SIMTIME'] = df[['PER','STP']].drop_duplicates().reset_index().set_index(['PER','STP'])['index'].add(1).reindex(pd.MultiIndex.from_frame(df[['PER','STP']])).to_numpy(dtype=float)
    meta = {'source_format':'binary_sfr2_istcb2_cell_array', 'headers': headers, 'zonebudget_ready': False}
    df.attrs.update(meta)
    return meta, df


def read_binary_rbr_records(raw: bytes, nfloat: int = 11, precision: str = 'auto', endian: str = '<', fortran: bool = False) -> pd.DataFrame:
    """Read a simple binary equivalent of the formatted listing.

    Expected per-row payload: PER, STP, LAYER, ROW, COL, SEG, RCH as int32,
    followed by 11 or 12 reals. This is intended as a testable binary pathway.
    Some compilers wrap unformatted records in 4-byte Fortran record markers;
    set fortran=True for one row per record.
    """
    if precision == 'auto':
        candidates = ['d','f']
    elif precision in ('double','float64','d'):
        candidates = ['d']
    elif precision in ('single','float32','f'):
        candidates = ['f']
    else:
        raise ValueError('precision must be auto, double, or single')
    last_err = None
    for code in candidates:
        real_size = struct.calcsize(code)
        payload_size = 7*4 + nfloat*real_size
        fmt = struct.Struct(endian + '7i' + code*nfloat)
        recs = []
        try:
            if fortran:
                off = 0
                while off < len(raw):
                    if off + 4 > len(raw): raise ValueError('trailing bytes')
                    reclen = struct.unpack(endian+'i', raw[off:off+4])[0]
                    off += 4
                    if reclen != payload_size: raise ValueError(f'record marker {reclen} != {payload_size}')
                    vals = fmt.unpack(raw[off:off+payload_size]); off += payload_size
                    endlen = struct.unpack(endian+'i', raw[off:off+4])[0]; off += 4
                    if endlen != reclen: raise ValueError('mismatched record marker')
                    recs.append(vals)
            else:
                if len(raw) % payload_size:
                    raise ValueError(f'file size {len(raw)} not multiple of record size {payload_size}')
                for off in range(0, len(raw), payload_size):
                    recs.append(fmt.unpack(raw[off:off+payload_size]))
            rows = []
            cols = LISTING_COLS_11 if nfloat == 11 else LISTING_COLS_12
            for vals in recs:
                per, stp, ilay, row, col, seg, rch = vals[:7]
                nums = vals[7:]
                rec = {'PER':per,'STP':stp,'LAYER':ilay,'ROW':row,'COL':col,'SEG':seg,'RCH':rch}
                rec.update(dict(zip(cols[5:], nums)))
                rows.append(rec)
            df = pd.DataFrame(rows)
            df['DELT'] = 1.0
            keys = df[['PER','STP']].drop_duplicates().reset_index(drop=True)
            key_to_i = {(int(p), int(s)): i+1 for i, (p, s) in enumerate(keys.to_numpy().tolist())}
            df['SIMTIME'] = [float(key_to_i[(p,s)]) for p,s in zip(df.PER, df.STP)]
            df['DATE_START'] = ''
            df['DATE_TIME'] = pd.NaT
            return _normalize(df, source_format=f"binary_rbr_{'fortran_' if fortran else ''}{'double' if code=='d' else 'single'}")
        except Exception as e:
            last_err = e
            continue
    raise ValueError(f'Could not parse binary RBR records: {last_err}')


def _normalize(df: pd.DataFrame, source_format: str) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    aliases = {
        'SEG.NO.':'SEG', 'RCH.':'RCH', 'FLOW INTO STRM. RCH.':'FLOW_IN',
        'FLOW TO AQUIFER':'FLOW_SEEPAGE', 'FLOW OUT OF STRM. RCH.':'FLOW_OUT',
        'OVRLND. RUNOFF':'RUNOFF', 'DIRECT PRECIP':'PRECIP', 'STREAM ET':'STREAM_ET',
        'STREAM HEAD':'HEAD_STREAM', 'STREAM DEPTH':'DEPTH_STREAM', 'STREAM WIDTH':'WIDTH_STREAM',
        'STREAMBED CONDCTNC.':'COND_STREAM', 'STREAMBED GRADIENT':'HEAD_GRADIENT',
        'STREAMBED ELEVATION':'ELEV_UP_STREAM'
    }
    df = df.rename(columns={k:v for k,v in aliases.items() if k in df.columns})
    for c in ['PER','STP','LAYER','ROW','COL','SEG','RCH']:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors='coerce').astype('Int64')
    for c in [x for x in df.columns if x not in ('DATE_START','DATE_TIME') and x not in ['PER','STP','LAYER','ROW','COL','SEG','RCH']]:
        df[c] = pd.to_numeric(df[c], errors='coerce')
    if 'DATE_START' in df.columns and 'DATE_TIME' not in df.columns:
        df['DATE_TIME'] = pd.to_datetime(df['DATE_START'].astype(str).str.replace('T',' ', regex=False), errors='coerce')
    for c in ['DELT','SIMTIME']:
        if c not in df.columns:
            df[c] = np.nan
    missing = [c for c in REQ if c not in df.columns]
    if missing:
        raise ValueError(f'Missing required normalized columns: {missing}')
    df.attrs['source_format'] = source_format
    df.attrs['zonebudget_ready'] = True
    order = ['DATE_START','DATE_TIME','PER','STP','DELT','SIMTIME','LAYER','ROW','COL','SEG','RCH'] + DB_TAIL_COLS
    cols = [c for c in order if c in df.columns] + [c for c in df.columns if c not in order]
    return df[cols]


def read_sfr_output_any_bytes(raw: bytes, allow_cell_array: bool = False, **binary_kwargs) -> Tuple[str, pd.DataFrame]:
    if _looks_text(raw):
        text = _decode(raw)
        if 'STREAM LISTING' in text[:100000].upper():
            df = read_sfr_reachbyreach_ascii_listing(text)
        else:
            df = read_db_ascii(text)
        return df.attrs.get('source_format','ascii'), df
    try:
        df = read_binary_db(raw)
        return df.attrs.get('source_format','binary_db'), df
    except Exception:
        pass
    try:
        meta, df = read_sfr2_istcb2_negative_binary_array(raw)
        if allow_cell_array:
            return meta['source_format'], df
        texts = ', '.join(sorted(set(df['BINARY_TEXT'].astype(str))))
        raise SfrOutputNotZoneBudgetReady(
            "Detected SFR2-style ISTCB2<0 binary cell-array output "
            f"({texts}). This file is parseable, but it contains gridded model-cell "
            "values rather than the full reach-by-reach budget table required by "
            "SFR ZoneBudget. This is common for SFR2 with negative ISTCB2 and "
            "differs from some older SFR1 behavior. Rerun SFR with a positive "
            "ISTCB2 value to generate the formatted full reach-by-reach listing. "
            "DB-style reach output is also acceptable where available."
        )
    except SfrOutputNotZoneBudgetReady:
        raise
    except Exception:
        df = read_binary_rbr_records(raw, **binary_kwargs)
        return df.attrs.get('source_format','binary_rbr'), df


def _default_out_csv_path(input_path: str) -> str:
    base, _ = os.path.splitext(input_path)
    return base + '_normalized.csv'


def _read_binary_member(path: str, member_name: str = "") -> Tuple[str, bytes]:
    if path.lower().endswith(".zip"):
        with zipfile.ZipFile(path, "r") as zf:
            members = zf.namelist()
            if member_name and member_name in members:
                name = member_name
            else:
                name = members[0]
            return name, zf.read(name)
    else:
        with open(path, "rb") as f:
            return os.path.basename(path), f.read()


def read_sfr_output(path: str, zip_member: str = "") -> Tuple[str, str, pd.DataFrame]:
    """
    Returns (format, source_name, normalized dataframe).

    Supported formats:
      - OWHM DB-style reach table, ASCII or binary (including zipped members)
      - SFR ISTCB2 positive formatted reach-by-reach listing, ASCII
      - Experimental binary reach-by-reach records from the companion reader

    SFR2 negative-ISTCB2 binary cell-array output is auto-detected and rejected
    because it is not a full reach-by-reach budget table.
    """
    name, raw = _read_binary_member(path, zip_member)
    binary_kwargs = dict(
        nfloat=int(BINARY_RBR_NFLOAT),
        precision=str(BINARY_RBR_PRECISION),
        endian=str(BINARY_RBR_ENDIAN),
        fortran=bool(BINARY_RBR_FORTRAN),
    )
    fmt, df = read_sfr_output_any_bytes(
        raw,
        allow_cell_array=bool(ALLOW_SFR2_NEGATIVE_ISTCB2_CELL_ARRAY),
        **binary_kwargs,
    )
    return fmt, name, df


def build_zonebudget_excel(
    df: pd.DataFrame,
    routing: pd.DataFrame,
    zone_mode: str,
    zones: pd.DataFrame,
    out_xlsx: str,
    meta: Dict,
    lake_budget: Optional[pd.DataFrame] = None,
) -> None:
    # Build zone mapping
    if zone_mode == "segment":
        seg_to_zone = dict(zip(zones["segment"], zones["zone"]))
        df["ZONE"] = df["SEG"].map(seg_to_zone).fillna(0).astype(int)
        seg_zone = lambda s: seg_to_zone.get(int(s), 0)
        reach_zone = None
    else:
        key_to_zone = {(int(r.segment), int(r.reach)): int(r.zone) for r in zones.itertuples(index=False)}
        df["ZONE"] = [key_to_zone.get((int(s), int(r)), 0) for s, r in zip(df["SEG"].values, df["RCH"].values)]
        seg_zone = None
        reach_zone = lambda s, r: key_to_zone.get((int(s), int(r)), 0)

    # Lake zone mapping. Lakes can be assigned in by-segment zone files using
    # negative Segment values (e.g., Segment=-3 means Lake 3). By-reach zone files
    # cannot assign lake zones, so lake rows default to Zone 0 if LAK budgets are used.
    if zone_mode == "segment":
        lake_to_zone = {abs(int(seg)): int(zone) for seg, zone in zip(zones["segment"], zones["zone"]) if int(seg) < 0}
    else:
        lake_to_zone = {}

    if lake_budget is not None:
        lake_budget = lake_budget.copy()
        lake_budget["ZONE"] = lake_budget["LAKE"].map(lake_to_zone).fillna(0).astype(int)

    # Zones list (include 0, SFR zones, and any lake-only zones)
    lake_zones = set(lake_budget["ZONE"].unique().tolist()) if lake_budget is not None else set()
    zones_list = sorted(set(df["ZONE"].unique().tolist()) | lake_zones | {0})

    # Timestep keys
    tkeys = df[["DATE_TIME", "PER", "STP", "DELT", "SIMTIME"]].drop_duplicates().sort_values(["PER", "STP"]).reset_index(drop=True)

    # Segment-level routing dicts
    routing = routing.copy()
    routing["segment"] = routing["segment"].astype(int)
    seg_outseg = dict(zip(routing["segment"], routing["outseg_norm"].astype(int)))
    seg_iupseg = dict(zip(routing["segment"], routing["iupseg_norm"].astype(int)))
    seg_is_outmodel = dict(zip(routing["segment"], routing["is_out_of_model"].astype(bool)))
    seg_is_head = dict(zip(routing["segment"], routing["is_head_segment"].astype(bool)))

    # Diversion DEST mapping using IUPSEG: for each destination D with iupseg=S => S->D
    src_to_dests: Dict[int, List[int]] = {}
    for r in routing.itertuples(index=False):
        s = int(r.iupseg_norm)
        d = int(r.segment)
        if s > 0:
            src_to_dests.setdefault(s, []).append(d)

    # Segment inflow/outflow per timestep (use reach ordering)
    grp = ["PER", "STP", "SEG"]
    seg_in = df.sort_values(grp + ["RCH"]).groupby(grp, as_index=False).first()[["PER", "STP", "SEG", "FLOW_IN"]].rename(columns={"FLOW_IN": "SEG_FLOW_IN"})
    seg_out = df.sort_values(grp + ["RCH"]).groupby(grp, as_index=False).last()[["PER", "STP", "SEG", "FLOW_OUT"]].rename(columns={"FLOW_OUT": "SEG_FLOW_OUT"})
    seg_flow = seg_in.merge(seg_out, on=grp, how="outer").fillna(0.0)

    # Segment zones
    if zone_mode == "segment":
        seg_flow["ZONE"] = seg_flow["SEG"].map(seg_to_zone).fillna(0).astype(int)
    else:
        # approximate segment zone as the most common reach zone in that segment (timestep-invariant)
        seg_zone_mode = df.groupby("SEG")["ZONE"].agg(lambda s: int(s.value_counts().idxmax())).to_dict()
        seg_flow["ZONE"] = seg_flow["SEG"].map(seg_zone_mode).fillna(0).astype(int)

    # Add routing columns
    seg_flow["OUTSEG"] = seg_flow["SEG"].map(seg_outseg).fillna(0).astype(int)
    seg_flow["IS_OUTMODEL"] = seg_flow["SEG"].map(seg_is_outmodel).fillna(True).astype(bool)

    # Inbound from upstream routing (downstream edges)
    up_to_down = seg_flow[["PER", "STP", "SEG", "OUTSEG", "SEG_FLOW_OUT"]].copy()
    up_to_down = up_to_down[up_to_down["OUTSEG"] > 0]
    inbound = up_to_down.groupby(["PER", "STP", "OUTSEG"], as_index=False)["SEG_FLOW_OUT"].sum().rename(columns={"OUTSEG": "SEG", "SEG_FLOW_OUT": "INBOUND_FROM_UPSTREAM"})
    seg_flow = seg_flow.merge(inbound, on=["PER", "STP", "SEG"], how="left")
    seg_flow["INBOUND_FROM_UPSTREAM"] = seg_flow["INBOUND_FROM_UPSTREAM"].fillna(0.0)    # ---------------- Diversions ----------------
    # Two diversion mechanisms are handled:
    #
    # (A) SFR-defined diversions (defined in the SFR input file):
    #     These are not reported as outflow from the source segment. Instead, they appear as FLOW_IN
    #     to the diversion (destination) segment. The diversion source is given by IUPSEG for the
    #     destination segment.
    #     For diversion destination segments (IUPSEG>0), we estimate diverted inflow as:
    #         DIV_SFR_TO_DEST = max(0, SEG_FLOW_IN(dest) - INBOUND_FROM_UPSTREAM(dest))
    #
    # (B) Farm Process semi-routed diversions:
    #     These may be indicated by negative RUNOFF at the diversion SOURCE reach/segment.
    #     Because RUNOFF can include natural runoff (+) and diversion (-), this is a NET indicator:
    #         DIV_FARM_NET_SRC = max(0, -RUNOFF_reported)
    #
    # For zone-boundary accounting we compute:
    #   - Cross-zone diversion transfers from (A) using IUPSEG mapping.
    #   - Additional cross-zone diversion transfers from (B) as residual when not already represented by (A).

    # --- (B) Farm net diversion indicator at SOURCE segment ---
    df["DIV_FARM_NET_SRC_REACH"] = np.where(df["RUNOFF"] < 0, -df["RUNOFF"], 0.0)
    div_farm_src_seg = df.groupby(["PER", "STP", "SEG"], as_index=False)["DIV_FARM_NET_SRC_REACH"].sum().rename(
        columns={"DIV_FARM_NET_SRC_REACH": "DIV_FARM_NET_SRC_SEG"}
    )
    seg_flow = seg_flow.merge(div_farm_src_seg, on=["PER", "STP", "SEG"], how="left")
    seg_flow["DIV_FARM_NET_SRC_SEG"] = seg_flow["DIV_FARM_NET_SRC_SEG"].fillna(0.0)

    # --- (A) SFR-defined diversion inflow to destination segments (IUPSEG>0) ---
    seg_flow["IUPSEG"] = seg_flow["SEG"].map(seg_iupseg).fillna(0).astype(int)
    is_div_dest = seg_flow["IUPSEG"] > 0
    seg_flow["DIV_SFR_TO_DEST"] = 0.0
    seg_flow.loc[is_div_dest, "DIV_SFR_TO_DEST"] = (seg_flow.loc[is_div_dest, "SEG_FLOW_IN"] - seg_flow.loc[is_div_dest, "INBOUND_FROM_UPSTREAM"]).clip(lower=0.0)

    # Transfer table (SFR-defined): source=IUPSEG(dest) -> dest=SEG
    div_sfr_tr = seg_flow[is_div_dest & (seg_flow["DIV_SFR_TO_DEST"] > 0)][
        ["PER", "STP", "IUPSEG", "SEG", "DIV_SFR_TO_DEST"]
    ].copy()
    div_sfr_tr = div_sfr_tr.rename(columns={"IUPSEG": "FROM_SEG", "SEG": "TO_SEG", "DIV_SFR_TO_DEST": "Q"})
    if div_sfr_tr.empty:
        div_sfr_tr = pd.DataFrame(columns=["PER", "STP", "FROM_SEG", "TO_SEG", "Q"])

    # Totals per diversion SOURCE segment (used to avoid double-counting in downstream transfers)
    div_sfr_from = div_sfr_tr.groupby(["PER", "STP", "FROM_SEG"], as_index=False)["Q"].sum().rename(
        columns={"FROM_SEG": "SEG", "Q": "DIV_SFR_FROM_SOURCE"}
    )
    seg_flow = seg_flow.merge(div_sfr_from, on=["PER", "STP", "SEG"], how="left")
    seg_flow["DIV_SFR_FROM_SOURCE"] = seg_flow["DIV_SFR_FROM_SOURCE"].fillna(0.0)


    # Totals per source represented by SFR-defined diversions
    sfr_by_source = div_sfr_tr.groupby(["PER", "STP", "FROM_SEG"], as_index=False)["Q"].sum().rename(columns={"Q": "DIV_SFR_FROM_SOURCE"})

    # --- Residual FARM diversions ---
    # NOTE: FMP semi-routed diversions (negative RUNOFF) are treated as OUTFLOW from the stream system
    # and are NOT included in interzone transfers (they should never appear in IN_FROM_ZONE_* / OUT_TO_ZONE_*).
    #
    # Therefore, interzone diversion transfers include ONLY SFR-defined diversions from div_sfr_tr.
    div_tr = div_sfr_tr.copy()
    if div_tr.empty:
        div_tr = pd.DataFrame(columns=["PER", "STP", "FROM_SEG", "TO_SEG", "Q"])

    # Determine zone of segments
    if zone_mode == "segment":
        div_tr["FROM_ZONE"] = div_tr["FROM_SEG"].map(seg_to_zone).fillna(0).astype(int)
        div_tr["TO_ZONE"] = div_tr["TO_SEG"].map(seg_to_zone).fillna(0).astype(int)
    else:
        # use seg_zone_mode computed above
        div_tr["FROM_ZONE"] = div_tr["FROM_SEG"].map(seg_zone_mode).fillna(0).astype(int)
        div_tr["TO_ZONE"] = div_tr["TO_SEG"].map(seg_zone_mode).fillna(0).astype(int)

    # Split into internal diversion QC vs interzone diversion transfer
    internal_div = div_tr[div_tr["FROM_ZONE"] == div_tr["TO_ZONE"]].groupby(["PER", "STP", "FROM_ZONE"], as_index=False)["Q"].sum()
    internal_div = internal_div.rename(columns={"FROM_ZONE": "ZONE", "Q": "DIVERSION_INTERNAL_QC"})
    div_xzone = div_tr[div_tr["FROM_ZONE"] != div_tr["TO_ZONE"]].copy()

    # Downstream interzone transfers (segment outflow)
    # If segment routes to OUTSEG within model and zones differ, transfer = SEG_FLOW_OUT
    if zone_mode == "segment":
        out_zone = seg_flow["OUTSEG"].map(seg_to_zone).fillna(0).astype(int)
    else:
        out_zone = seg_flow["OUTSEG"].map(seg_zone_mode).fillna(0).astype(int)
    seg_flow["OUT_ZONE"] = out_zone

    down_tr = seg_flow[(seg_flow["OUTSEG"] > 0) & (seg_flow["ZONE"] != seg_flow["OUT_ZONE"])][
        ["PER", "STP", "SEG", "OUTSEG", "SEG_FLOW_OUT", "DIV_SFR_FROM_SOURCE", "ZONE", "OUT_ZONE"]
    ].copy()
    # IMPORTANT: SEG_FLOW_OUT at diversion sources can include water diverted to SFR diversion segments.
    # To avoid double-counting, subtract SFR-defined diversions taken from the source when computing
    # downstream interzone transfers.
    down_tr["Q"] = (down_tr["SEG_FLOW_OUT"] - down_tr["DIV_SFR_FROM_SOURCE"].fillna(0.0)).clip(lower=0.0)
    down_tr = down_tr.rename(columns={"SEG": "FROM_SEG", "OUTSEG": "TO_SEG", "ZONE": "FROM_ZONE", "OUT_ZONE": "TO_ZONE"})
    down_tr["TYPE"] = "DOWNSTREAM"

    div_xzone["TYPE"] = "DIVERSION_NET"

    # ---------------- Lake/stream transfers ----------------
    # SFR-to-lake transfers are explicit in SFR routing as OUTSEG<0; the transfer
    # amount is the source segment FLOW_OUT. Lake-to-SFR transfers are represented
    # by destination segments with IUPSEG<0; the amount is estimated from residual
    # segment inflow after routed upstream inflow and SFR-defined diversion inflow.
    lake_tr_parts = []

    sfr_to_lake = seg_flow[seg_flow["OUTSEG"] < 0][["PER", "STP", "SEG", "OUTSEG", "SEG_FLOW_OUT", "ZONE"]].copy()
    if not sfr_to_lake.empty:
        sfr_to_lake["LAKE"] = sfr_to_lake["OUTSEG"].abs().astype(int)
        sfr_to_lake["FROM_ZONE"] = sfr_to_lake["ZONE"].astype(int)
        sfr_to_lake["TO_ZONE"] = sfr_to_lake["LAKE"].map(lake_to_zone).fillna(0).astype(int)
        sfr_to_lake["Q"] = sfr_to_lake["SEG_FLOW_OUT"].clip(lower=0.0)
        sfr_to_lake["TYPE"] = "SFR_TO_LAKE"
        lake_tr_parts.append(sfr_to_lake[["PER", "STP", "FROM_ZONE", "TO_ZONE", "Q", "TYPE"]])

    is_lake_inflow_seg = seg_flow["IUPSEG"] < 0
    seg_flow["LAKE_TO_SFR_IN_SEG"] = 0.0
    seg_flow.loc[is_lake_inflow_seg, "LAKE_TO_SFR_IN_SEG"] = (
        seg_flow.loc[is_lake_inflow_seg, "SEG_FLOW_IN"]
        - seg_flow.loc[is_lake_inflow_seg, "INBOUND_FROM_UPSTREAM"]
        - seg_flow.loc[is_lake_inflow_seg, "DIV_SFR_TO_DEST"].fillna(0.0)
    ).clip(lower=0.0)

    lake_to_sfr = seg_flow[is_lake_inflow_seg & (seg_flow["LAKE_TO_SFR_IN_SEG"] > 0)][
        ["PER", "STP", "SEG", "IUPSEG", "LAKE_TO_SFR_IN_SEG", "ZONE"]
    ].copy()
    if not lake_to_sfr.empty:
        lake_to_sfr["LAKE"] = lake_to_sfr["IUPSEG"].abs().astype(int)
        lake_to_sfr["FROM_ZONE"] = lake_to_sfr["LAKE"].map(lake_to_zone).fillna(0).astype(int)
        lake_to_sfr["TO_ZONE"] = lake_to_sfr["ZONE"].astype(int)
        lake_to_sfr["Q"] = lake_to_sfr["LAKE_TO_SFR_IN_SEG"].clip(lower=0.0)
        lake_to_sfr["TYPE"] = "LAKE_TO_SFR"
        lake_tr_parts.append(lake_to_sfr[["PER", "STP", "FROM_ZONE", "TO_ZONE", "Q", "TYPE"]])

    if lake_tr_parts:
        lake_tr = pd.concat(lake_tr_parts, ignore_index=True)
    else:
        lake_tr = pd.DataFrame(columns=["PER", "STP", "FROM_ZONE", "TO_ZONE", "Q", "TYPE"])

    lake_internal = lake_tr[lake_tr["FROM_ZONE"] == lake_tr["TO_ZONE"]].groupby(["PER", "STP", "FROM_ZONE"], as_index=False)["Q"].sum()
    lake_internal = lake_internal.rename(columns={"FROM_ZONE": "ZONE", "Q": "LAKE_STREAM_INTERNAL_QC"})
    lake_xzone = lake_tr[lake_tr["FROM_ZONE"] != lake_tr["TO_ZONE"]].copy()

    # Combine transfers (for IN/OUT columns)
    all_tr = pd.concat([
        down_tr[["PER","STP","FROM_ZONE","TO_ZONE","Q","TYPE"]],
        div_xzone[["PER","STP","FROM_ZONE","TO_ZONE","Q","TYPE"]],
        lake_xzone[["PER","STP","FROM_ZONE","TO_ZONE","Q","TYPE"]],
    ], ignore_index=True)
    tr_sum = all_tr.groupby(["PER","STP","FROM_ZONE","TO_ZONE"], as_index=False)["Q"].sum()

    # Approx diverted inflow to destination segments for headwater separation:
    # sum of allocated Q for each TO_SEG
    div_in_seg = div_tr.groupby(["PER","STP","TO_SEG"], as_index=False)["Q"].sum().rename(columns={"TO_SEG":"SEG","Q":"DIV_NET_IN_SEG"})
    seg_flow = seg_flow.merge(div_in_seg, on=["PER","STP","SEG"], how="left")
    seg_flow["DIV_NET_IN_SEG"] = seg_flow["DIV_NET_IN_SEG"].fillna(0.0)

    # Headwater external flow per segment: inflow - inbound_from_upstream - diverted_in (net indicator)
    if "LAKE_TO_SFR_IN_SEG" not in seg_flow.columns:
        seg_flow["LAKE_TO_SFR_IN_SEG"] = 0.0
    seg_flow["PRESCRIBED_INFLOW_RESIDUAL"] = seg_flow["SEG_FLOW_IN"] - seg_flow["INBOUND_FROM_UPSTREAM"] - seg_flow["DIV_NET_IN_SEG"] - seg_flow["LAKE_TO_SFR_IN_SEG"]
    seg_flow["PRESCRIBED_INFLOW_RESIDUAL"] = seg_flow["PRESCRIBED_INFLOW_RESIDUAL"].where(
        seg_flow["PRESCRIBED_INFLOW_RESIDUAL"] > float(ADDITIONAL_INFLOW_TOLERANCE), 0.0
    )
    seg_flow["IS_HEAD_SEGMENT"] = seg_flow["SEG"].map(seg_is_head).fillna(False).astype(bool)

    # FLOW_HEAD is prescribed inflow at true headwater segments. If a segment also
    # receives routed upstream flow, any residual inflow at that segment is reported
    # separately as ADDITIONAL_INFLOW. This captures non-head FLOW inputs, including
    # tabfile-driven FLOW values, without trying to read the SFR package/tabfiles.
    seg_flow["HEAD_EXT"] = np.where(seg_flow["IS_HEAD_SEGMENT"], seg_flow["PRESCRIBED_INFLOW_RESIDUAL"], 0.0)
    seg_flow["ADDITIONAL_INFLOW"] = np.where(~seg_flow["IS_HEAD_SEGMENT"], seg_flow["PRESCRIBED_INFLOW_RESIDUAL"], 0.0)

    # Reach-level terms by zone
    df["RUNOFF_POS"] = df["RUNOFF"].where(df["RUNOFF"] > 0, 0.0)
    df["FARM_DIV_OUT_REACH"] = np.where(df["RUNOFF"] < 0, -df["RUNOFF"], 0.0)
    zone_reach_terms = df.groupby(["PER","STP","ZONE"], as_index=False).agg(
        FLOW_SEEPAGE=("FLOW_SEEPAGE","sum"),
        RUNOFF=("RUNOFF_POS","sum"),
        PRECIP=("PRECIP","sum"),
        STREAM_ET=("STREAM_ET","sum"),
    )
    farm_div_zone = df.groupby(["PER","STP","ZONE"], as_index=False)["FARM_DIV_OUT_REACH"].sum().rename(columns={"FARM_DIV_OUT_REACH":"FARM_DIVERSION_NET_QC"})


    zone_head = seg_flow.groupby(["PER","STP","ZONE"], as_index=False)["HEAD_EXT"].sum().rename(columns={"HEAD_EXT":"FLOW_HEAD"})
    zone_additional = seg_flow.groupby(["PER","STP","ZONE"], as_index=False)["ADDITIONAL_INFLOW"].sum()

    zone_outmodel = seg_flow[seg_flow["IS_OUTMODEL"]].groupby(["PER","STP","ZONE"], as_index=False)["SEG_FLOW_OUT"].sum().rename(columns={"SEG_FLOW_OUT":"FLOW_OUTOFMODEL"})

    lake_cols = [
        "LAK_PRECIP", "LAK_EVAP", "LAK_RUNOFF", "LAK_GW_INFLOW", "LAK_GW_OUTFLOW",
        "LAK_WATER_USE", "LAK_STORAGE_CHANGE", "LAK_CONNECTED_LAKE_INFLUX_QC",
        "LAK_SW_INFLOW_QC", "LAK_SW_OUTFLOW_QC",
    ]
    if lake_budget is not None:
        zone_lake_terms = lake_budget.groupby(["PER", "STP", "ZONE"], as_index=False)[lake_cols].sum()
    else:
        zone_lake_terms = pd.DataFrame(columns=["PER", "STP", "ZONE"] + lake_cols)

    # helper to build per-zone timeseries
    def zone_ts(z: int) -> pd.DataFrame:
        ts = tkeys.copy()
        ts["ZONE"] = z

        sub = zone_reach_terms[zone_reach_terms["ZONE"] == z].drop(columns=["ZONE"])
        ts = ts.merge(sub, on=["PER","STP"], how="left")
        fd = farm_div_zone[farm_div_zone["ZONE"] == z].drop(columns=["ZONE"])
        ts = ts.merge(fd, on=["PER","STP"], how="left")

        zh = zone_head[zone_head["ZONE"] == z].drop(columns=["ZONE"])
        ts = ts.merge(zh, on=["PER","STP"], how="left")

        za = zone_additional[zone_additional["ZONE"] == z].drop(columns=["ZONE"])
        ts = ts.merge(za, on=["PER","STP"], how="left")

        zom = zone_outmodel[zone_outmodel["ZONE"] == z].drop(columns=["ZONE"])
        ts = ts.merge(zom, on=["PER","STP"], how="left")

        idv = internal_div[internal_div["ZONE"] == z].drop(columns=["ZONE"])
        ts = ts.merge(idv, on=["PER","STP"], how="left")

        lint = lake_internal[lake_internal["ZONE"] == z].drop(columns=["ZONE"])
        ts = ts.merge(lint, on=["PER","STP"], how="left")

        lz = zone_lake_terms[zone_lake_terms["ZONE"] == z].drop(columns=["ZONE"])
        ts = ts.merge(lz, on=["PER","STP"], how="left")

        for c in ["FLOW_SEEPAGE","RUNOFF","FARM_DIVERSION_NET_QC","PRECIP","STREAM_ET","FLOW_HEAD","ADDITIONAL_INFLOW","FLOW_OUTOFMODEL","DIVERSION_INTERNAL_QC","LAKE_STREAM_INTERNAL_QC"] + lake_cols:
            if c not in ts.columns:
                ts[c] = 0.0
            ts[c] = ts[c].fillna(0.0)

        for k in zones_list:
            ts[f"IN_FROM_ZONE_{k}"] = 0.0
            ts[f"OUT_TO_ZONE_{k}"] = 0.0

        inz = tr_sum[tr_sum["TO_ZONE"] == z]
        for _, r in inz.iterrows():
            ts.loc[(ts["PER"] == r["PER"]) & (ts["STP"] == r["STP"]), f"IN_FROM_ZONE_{int(r['FROM_ZONE'])}"] = float(r["Q"])

        outz = tr_sum[tr_sum["FROM_ZONE"] == z]
        for _, r in outz.iterrows():
            ts.loc[(ts["PER"] == r["PER"]) & (ts["STP"] == r["STP"]), f"OUT_TO_ZONE_{int(r['TO_ZONE'])}"] = float(r["Q"])

        in_cols = [f"IN_FROM_ZONE_{k}" for k in zones_list]
        out_cols = [f"OUT_TO_ZONE_{k}" for k in zones_list]
        ts["MASS_BALANCE_RESIDUAL"] = (
            ts["FLOW_HEAD"] + ts["ADDITIONAL_INFLOW"] + ts["RUNOFF"] + ts["PRECIP"]
            + ts["LAK_PRECIP"] + ts["LAK_RUNOFF"] + ts["LAK_GW_INFLOW"]
            + ts[in_cols].sum(axis=1)
            - (
                ts["FLOW_SEEPAGE"] + ts["STREAM_ET"] + ts["FLOW_OUTOFMODEL"] + ts["FARM_DIVERSION_NET_QC"]
                + ts["LAK_EVAP"] + ts["LAK_GW_OUTFLOW"] + ts["LAK_WATER_USE"] + ts["LAK_STORAGE_CHANGE"]
                + ts[out_cols].sum(axis=1)
            )
        )

        core = ["DATE_TIME","PER","STP","DELT","SIMTIME","ZONE",
                "FLOW_HEAD","ADDITIONAL_INFLOW","FLOW_SEEPAGE","FLOW_OUTOFMODEL","RUNOFF","FARM_DIVERSION_NET_QC","DIVERSION_INTERNAL_QC","PRECIP","STREAM_ET",
                "LAK_PRECIP","LAK_EVAP","LAK_RUNOFF","LAK_GW_INFLOW","LAK_GW_OUTFLOW","LAK_WATER_USE","LAK_STORAGE_CHANGE","LAKE_STREAM_INTERNAL_QC","LAK_CONNECTED_LAKE_INFLUX_QC","LAK_SW_INFLOW_QC","LAK_SW_OUTFLOW_QC"]
        inter = []
        for k in zones_list:
            inter += [f"IN_FROM_ZONE_{k}", f"OUT_TO_ZONE_{k}"]
        final_cols = core + inter + ["MASS_BALANCE_RESIDUAL"]
        return ts[final_cols]

    zone_tabs = {z: zone_ts(z) for z in zones_list}

    # ===== System-wide SFR balance (all zones combined) =====
    # Interzone transfers should cancel at the system scale; we report a total-balance tab for comparison
    # against groundwater budget SFR leakage term.
    sys_ts = tkeys.copy()
    sys_ts["SFR_SYSTEM"] = "ALL"

    # Sum reach-based terms across all zones
    sys_reach = df.groupby(["PER","STP"], as_index=False).agg(
        FLOW_SEEPAGE=("FLOW_SEEPAGE","sum"),
        RUNOFF=("RUNOFF_POS","sum"),
        PRECIP=("PRECIP","sum"),
        STREAM_ET=("STREAM_ET","sum"),
    )
    sys_ts = sys_ts.merge(sys_reach, on=["PER","STP"], how="left")

    # Sum headwater external inflow across all segments
    sys_head = seg_flow.groupby(["PER","STP"], as_index=False)["HEAD_EXT"].sum().rename(columns={"HEAD_EXT":"FLOW_HEAD"})
    sys_ts = sys_ts.merge(sys_head, on=["PER","STP"], how="left")

    sys_additional = seg_flow.groupby(["PER","STP"], as_index=False)["ADDITIONAL_INFLOW"].sum()
    sys_ts = sys_ts.merge(sys_additional, on=["PER","STP"], how="left")

    # Out-of-model outflow across all segments
    sys_out = seg_flow[seg_flow["IS_OUTMODEL"]].groupby(["PER","STP"], as_index=False)["SEG_FLOW_OUT"].sum().rename(columns={"SEG_FLOW_OUT":"FLOW_OUTOFMODEL"})
    sys_ts = sys_ts.merge(sys_out, on=["PER","STP"], how="left")

    # Total net diversion indicator across all diversion sources (QC)
    sys_div = df.groupby(["PER","STP"], as_index=False)["FARM_DIV_OUT_REACH"].sum().rename(columns={"FARM_DIV_OUT_REACH":"FARM_DIVERSION_NET_TOTAL_QC"})
    sys_ts = sys_ts.merge(sys_div, on=["PER","STP"], how="left")

    # Interzone transfer total (QC only; this is sum of all cross-zone transfers and does not enter the system balance)
    sys_xfer = tr_sum.groupby(["PER","STP"], as_index=False)["Q"].sum().rename(columns={"Q":"INTERZONE_TOTAL_QC"})
    sys_ts = sys_ts.merge(sys_xfer, on=["PER","STP"], how="left")

    if lake_budget is not None:
        sys_lake = lake_budget.groupby(["PER", "STP"], as_index=False)[lake_cols].sum()
    else:
        sys_lake = pd.DataFrame(columns=["PER", "STP"] + lake_cols)
    sys_ts = sys_ts.merge(sys_lake, on=["PER", "STP"], how="left")

    for c in ["FLOW_SEEPAGE","RUNOFF","PRECIP","STREAM_ET","FLOW_HEAD","ADDITIONAL_INFLOW","FLOW_OUTOFMODEL","FARM_DIVERSION_NET_TOTAL_QC","INTERZONE_TOTAL_QC"] + lake_cols:
        if c not in sys_ts.columns:
            sys_ts[c] = 0.0
        sys_ts[c] = sys_ts[c].fillna(0.0)

    # System mass balance residual (no interzone terms; they cancel at system scale)
    sys_ts["MASS_BALANCE_RESIDUAL_SYSTEM"] = (
        sys_ts["FLOW_HEAD"] + sys_ts["ADDITIONAL_INFLOW"] + sys_ts["RUNOFF"] + sys_ts["PRECIP"]
        + sys_ts["LAK_PRECIP"] + sys_ts["LAK_RUNOFF"] + sys_ts["LAK_GW_INFLOW"]
        - (
            sys_ts["FLOW_SEEPAGE"] + sys_ts["STREAM_ET"] + sys_ts["FLOW_OUTOFMODEL"] + sys_ts["FARM_DIVERSION_NET_TOTAL_QC"]
            + sys_ts["LAK_EVAP"] + sys_ts["LAK_GW_OUTFLOW"] + sys_ts["LAK_WATER_USE"] + sys_ts["LAK_STORAGE_CHANGE"]
        )
    )
    sys_cols = [
        "DATE_TIME","PER","STP","DELT","SIMTIME","SFR_SYSTEM",
        "FLOW_HEAD","ADDITIONAL_INFLOW","FLOW_SEEPAGE","FLOW_OUTOFMODEL","RUNOFF",
        "FARM_DIVERSION_NET_TOTAL_QC","PRECIP","STREAM_ET",
        "LAK_PRECIP","LAK_EVAP","LAK_RUNOFF","LAK_GW_INFLOW","LAK_GW_OUTFLOW","LAK_WATER_USE","LAK_STORAGE_CHANGE","LAK_CONNECTED_LAKE_INFLUX_QC","LAK_SW_INFLOW_QC","LAK_SW_OUTFLOW_QC",
        "INTERZONE_TOTAL_QC","MASS_BALANCE_RESIDUAL_SYSTEM"
    ]
    sys_ts = sys_ts[sys_cols]

    
    # =========================
    # Apply unit/time scaling
    # =========================
    # Compute model_volume -> output_volume factor (model_volume is in model_length^3)
    FT_PER_M = 3.280839895013123
    FT3_PER_M3 = FT_PER_M ** 3
    FT3_PER_ACFT = 43560.0  # 1 acre-foot = 43,560 ft^3

    if OUTPUT_VOLUME_UNIT.lower() == "custom":
        vol_factor = float(VOLUME_CONV_FACTOR)
    else:
        ml = MODEL_LENGTH_UNIT.lower()
        if ml == "ft":
            model_ft3_per_model_vol = 1.0
        elif ml == "m":
            model_ft3_per_model_vol = FT3_PER_M3
        elif ml == "custom":
            model_ft3_per_model_vol = float(FT_PER_MODEL_LEN) ** 3
        else:
            raise ValueError(f"Unrecognized MODEL_LENGTH_UNIT: {MODEL_LENGTH_UNIT}")

        outu = OUTPUT_VOLUME_UNIT.lower()
        if outu == "ft3":
            vol_factor = model_ft3_per_model_vol
        elif outu == "m3":
            vol_factor = model_ft3_per_model_vol / FT3_PER_M3
        elif outu in ("acft", "acre-ft", "acreft"):
            vol_factor = model_ft3_per_model_vol / FT3_PER_ACFT
        else:
            raise ValueError(f"Unrecognized OUTPUT_VOLUME_UNIT: {OUTPUT_VOLUME_UNIT}")

    def _scale_df(df_in: pd.DataFrame, is_system: bool = False) -> pd.DataFrame:
        df = df_in.copy()
        base_cols = [
            "FLOW_HEAD", "ADDITIONAL_INFLOW", "FLOW_SEEPAGE", "FLOW_OUTOFMODEL",
            "RUNOFF", "PRECIP", "STREAM_ET",
        ]
        extra_cols = [
            "FARM_DIVERSION_NET_QC",
            "DIVERSION_INTERNAL_QC",
            "FARM_DIVERSION_NET_TOTAL_QC",
            "INTERZONE_TOTAL_QC",
            "LAK_PRECIP", "LAK_EVAP", "LAK_RUNOFF", "LAK_GW_INFLOW", "LAK_GW_OUTFLOW",
            "LAK_WATER_USE", "LAK_STORAGE_CHANGE", "LAKE_STREAM_INTERNAL_QC",
            "LAK_CONNECTED_LAKE_INFLUX_QC", "LAK_SW_INFLOW_QC", "LAK_SW_OUTFLOW_QC",
        ]
        flux_cols = [c for c in base_cols + extra_cols if c in df.columns]
        inter_cols = [c for c in df.columns if c.startswith("IN_FROM_ZONE_") or c.startswith("OUT_TO_ZONE_")]

        all_scale_cols = flux_cols + inter_cols
        if is_system and "MASS_BALANCE_RESIDUAL_SYSTEM" in df.columns:
            all_scale_cols.append("MASS_BALANCE_RESIDUAL_SYSTEM")
        if (not is_system) and "MASS_BALANCE_RESIDUAL" in df.columns:
            all_scale_cols.append("MASS_BALANCE_RESIDUAL")

        if OUTPUT_BASIS.upper() == "PER_STRESS_PERIOD":
            if df["DELT"].isna().any():
                raise ValueError(
                    "OUTPUT_BASIS='PER_STRESS_PERIOD' requires DELT values, but the selected SFR output "
                    "does not include DELT for one or more timesteps. This is expected for ISTCB2 formatted "
                    "reach-by-reach listings. Use OUTPUT_BASIS='PER_DAY' to report rates, or use an output "
                    "format that includes DELT if integrated volume per stress period is needed."
                )
            factor = df["DELT"].astype(float) * vol_factor
        elif OUTPUT_BASIS.upper() == "PER_DAY":
            if MODEL_TIME_UNIT_IN_DAYS == 0:
                raise ValueError("MODEL_TIME_UNIT_IN_DAYS must be non-zero for OUTPUT_BASIS='PER_DAY'.")
            factor = pd.Series(vol_factor / MODEL_TIME_UNIT_IN_DAYS, index=df.index)
        else:
            raise ValueError(f"Unrecognized OUTPUT_BASIS: {OUTPUT_BASIS}")

        for c in all_scale_cols:
            df[c] = df[c].astype(float) * factor

        return df

    # Scale zone tabs
    for z in zone_tabs:
        zone_tabs[z] = _scale_df(zone_tabs[z], is_system=False)

    # Scale system tab
    sys_ts = _scale_df(sys_ts, is_system=True)

# Write workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "README_METADATA"

    meta_rows = [
        ("Generated", datetime.now().isoformat(timespec="seconds")),
        ("SFR input", meta.get("sfr_input","")),
        ("SFR output", meta.get("sfr_output","")),
        ("SFR output member", meta.get("sfr_output_member","")),
        ("SFR output format", meta.get("sfr_output_format","")),
        ("Zone config", meta.get("zone_config","")),
        ("Zone mode", zone_mode),
        ("Default zone for unspecified", 0),
        ("Routing CSV (QC)", meta.get("routing_csv","")),
        ("Routing edge CSV (QC)", meta.get("routing_edges_csv","")),
        ("Lake routing QC CSV", meta.get("lake_qc_csv","")),
        ("LAK input", meta.get("lak_input","")),
        ("LAK budget CSV", meta.get("lak_budget_csv","")),
        ("LAK budget row count", meta.get("lak_budget_rows","")),
        ("LAK NLAKES", meta.get("lak_nlakes","")),
        ("LAK NSLMS", meta.get("lak_nslms","")),
        ("LAK parser warnings", meta.get("lak_warnings","")),
        ("Lake/sublake note", "Negative SFR OUTSEG values are treated as SFR-to-lake edges; negative IUPSEG values are treated as lake-to-SFR edges. LAK sublake systems are written as lake-to-lake QC edges only; internal sublake flows are not quantified."),
        ("LAK budget handling note", "Optional LAK budget CSV terms are added to zone budgets using lake zones from negative Segment rows in the by-segment zone CSV. SFR-to-lake and lake-to-SFR transfers are mapped from SFR routing and SFR output; LAK SW_INFLOW/SW_OUTFLOW are retained as QC columns to avoid double-counting stream-lake transfers."),
        ("NSTRM", meta.get("nstrm","")),
        ("NSS", meta.get("nss","")),
        ("Timestep count", len(tkeys)),
        ("Zones present (including 0)", ", ".join(str(z) for z in zones_list)),
        ("Diversion method", "SFR-defined diversions stay in the stream system (accounted via IUPSEG). FMP semi-routed diversions (negative RUNOFF) leave the stream system and are treated as FARM_DIVERSION_NET."),
        ("Diversion caveat", "RUNOFF at diversion sources may include natural runoff (+) and diversion (-); FARM_DIVERSION_NET=max(0,-RUNOFF) is a net indicator, not guaranteed gross diversion."),
        ("Additional inflow method", "For non-head segments, ADDITIONAL_INFLOW=max(0, segment FLOW_IN - routed upstream inflow - SFR diversion inflow). This captures prescribed FLOW at non-head locations, including tabfile-driven FLOW, without reading tabfiles."),
        ("Additional inflow tolerance", ADDITIONAL_INFLOW_TOLERANCE),
        ("Output basis", OUTPUT_BASIS),
        ("Model length unit", MODEL_LENGTH_UNIT),
        ("Output volume unit", OUTPUT_VOLUME_UNIT),
        ("Custom volume factor", VOLUME_CONV_FACTOR if OUTPUT_VOLUME_UNIT.lower()=="custom" else ""),
                ("Model time unit in days", MODEL_TIME_UNIT_IN_DAYS),
        ("Volume units label", VOLUME_UNIT_LABEL),
        ("Notes", "DATE_START parsed to DATE_TIME after replacing 'T' with space. All reported flux terms are scaled to either integrated volume per stress period or average rate in volume-per-day, depending on OUTPUT_BASIS. For ISTCB2 formatted listings, DELT and SIMTIME are not present in the SFR output and are left blank. MASS_BALANCE_RESIDUAL is in the same units."),
    ]

    ws.append(["Field", "Value"])
    ws["A1"].font = Font(bold=True)
    ws["B1"].font = Font(bold=True)
    for k, v in meta_rows:
        ws.append([k, v])

    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 120

    # System-wide tab
    ws_sys = wb.create_sheet(title="SFR_TOTAL")
    for r in dataframe_to_rows(sys_ts, index=False, header=True):
        ws_sys.append(r)
    for cell in ws_sys[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws_sys.freeze_panes = "A2"
    ws_sys.auto_filter.ref = ws_sys.dimensions
    ws_sys.column_dimensions["A"].width = 20
    ws_sys.column_dimensions["B"].width = 6
    ws_sys.column_dimensions["C"].width = 6
    ws_sys.column_dimensions["D"].width = 10
    ws_sys.column_dimensions["E"].width = 12
    ws_sys.column_dimensions["F"].width = 12
    for col in range(7, ws_sys.max_column + 1):
        ws_sys.column_dimensions[ws_sys.cell(row=1, column=col).column_letter].width = 20

    for z, zdf in zone_tabs.items():
        title = f"Zone_{z}"[:31]
        wz = wb.create_sheet(title=title)
        for r in dataframe_to_rows(zdf, index=False, header=True):
            wz.append(r)
        for cell in wz[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        wz.freeze_panes = "A2"
        wz.auto_filter.ref = wz.dimensions
        # basic widths
        wz.column_dimensions["A"].width = 20
        wz.column_dimensions["B"].width = 6
        wz.column_dimensions["C"].width = 6
        wz.column_dimensions["D"].width = 10
        wz.column_dimensions["E"].width = 12
        wz.column_dimensions["F"].width = 6
        for col in range(7, wz.max_column + 1):
            wz.column_dimensions[wz.cell(row=1, column=col).column_letter].width = 16

    os.makedirs(os.path.dirname(out_xlsx) or ".", exist_ok=True)
    wb.save(out_xlsx)


def run():
    # Output paths
    routing_csv = OUT_ROUTING_CSV_PATH.strip()
    if not routing_csv:
        base = os.path.splitext(OUT_EXCEL_PATH)[0]
        routing_csv = base + "_routing.csv"

    # Parse routing and write QC CSV
    routing_res = parse_sfr_routing_table(SFR_INPUT_PATH)
    routing = routing_res["routing_table"].copy()

    lak_info = parse_lak_sublake_systems(LAK_INPUT_PATH)
    routing_edges_csv = OUT_ROUTING_EDGES_CSV_PATH.strip()
    if not routing_edges_csv:
        base, _ = os.path.splitext(routing_csv)
        routing_edges_csv = base + "_edges.csv"
    lake_qc_csv = os.path.splitext(routing_edges_csv)[0] + "_lake_qc.csv"

    routing.to_csv(routing_csv, index=False)
    routing_edges = build_routing_edges_table(routing, lak_info)
    routing_edges.to_csv(routing_edges_csv, index=False)
    lake_qc = build_lake_routing_qc(routing, lak_info)
    lake_qc.to_csv(lake_qc_csv, index=False)

    # Read zones
    zone_mode, zones = read_zone_config(ZONE_CONFIG_PATH)

    # Read SFR output
    fmt, member, df = read_sfr_output(SFR_OUTPUT_PATH, ZIP_MEMBER_NAME)

    # Read optional LAK budget CSV produced by the companion listing-file scraper
    lake_budget = read_lake_budget_csv(LAK_BUDGET_CSV_PATH)

    # Validate required columns
    needed = ["DATE_TIME","PER","STP","DELT","SIMTIME","SEG","RCH","FLOW_IN","FLOW_OUT","RUNOFF","PRECIP","STREAM_ET","FLOW_SEEPAGE"]
    missing = [c for c in needed if c not in df.columns]
    if missing:
        raise ValueError(f"SFR output missing required columns: {missing}")

    meta = dict(
        sfr_input=SFR_INPUT_PATH,
        sfr_output=SFR_OUTPUT_PATH,
        sfr_output_member=member,
        sfr_output_format=fmt,
        zone_config=ZONE_CONFIG_PATH,
        routing_csv=routing_csv,
        routing_edges_csv=routing_edges_csv,
        lake_qc_csv=lake_qc_csv,
        lak_input=LAK_INPUT_PATH,
        lak_budget_csv=LAK_BUDGET_CSV_PATH,
        lak_budget_rows=0 if lake_budget is None else len(lake_budget),
        lak_nlakes=lak_info.get("nlakes", ""),
        lak_nslms=lak_info.get("nslms", ""),
        lak_warnings=" | ".join(lak_info.get("warnings", [])),
        nstrm=routing_res["nstrm"],
        nss=routing_res["nss"],
    )

    build_zonebudget_excel(df, routing, zone_mode, zones, OUT_EXCEL_PATH, meta, lake_budget=lake_budget)

    print("=== SFR ZoneBudget ===")
    print(f"Wrote routing CSV:      {routing_csv}")
    print(f"Wrote routing edge CSV: {routing_edges_csv}")
    print(f"Wrote lake QC CSV:      {lake_qc_csv}")
    if lake_budget is not None:
        print(f"Read LAK budget rows:   {len(lake_budget)}")
    print(f"Wrote Excel:            {OUT_EXCEL_PATH}")


if __name__ == "__main__":
    run()
